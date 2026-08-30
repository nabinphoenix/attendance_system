import io,json,secrets
from dataclasses import dataclass
from datetime import UTC,date,datetime,time
from typing import Annotated
import pandas as pd
from fastapi import APIRouter,Depends,File,HTTPException,UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from jinja2 import Template
from sqlalchemy import String,cast,func,or_,select
from app.core.config import settings
from app.core.dependencies import DbSession,get_current_user,require_role
from app.core.security import hash_password
from app.modules.academic.models import AcademicModule,Batch,Block,ClassType,Guardian,Intake,ModuleOffering,Room,RoutineEntry,RoutineEntrySection,RoutinePendingSection,Section,Student,Subject,Teacher,TimeSlot
from app.modules.academic.invitation_service import issue_student_invitation
from app.modules.academic.module_offering_service import offering_section_ids
from app.modules.academic.routine_router import RoutineCreate,check_routine_conflicts,create_or_merge_routine_entry,create_routine_entry,matching_physical_routine,merge_routine_entry_sections,persist_pending_section_references,valid_routine
from app.modules.attendance.models import AttendanceRecord,AttendanceStatus
from app.modules.course_completion.models import CoursePlan
from app.modules.crm.models import CaseStatus,StudentCase
from app.modules.identity.models import User,UserRole
from app.modules.academic.student_profile_service import current_student_profile
from app.modules.scheduling.models import ClassSession,TimetableEntry
from .models import AuditLog,ImportJob,Notification,NotificationStatus
from .schemas import AuditPage,AuditRead,ImportError,ImportJobRead,NotificationRead
from .service import log_audit
router=APIRouter(tags=["operations"])

TEACHER_TIMETABLE_COLUMNS=["intake_code","semester","sections","day","start_time","end_time","module_code","class_type","block","room"]
SECTION_ROUTINE_COLUMNS=["day","start_time","end_time","sections","module_code","module_title","class_type","lecturer_email","block","room"]
DAY_CODES={"mon":0,"monday":0,"tue":1,"tuesday":1,"wed":2,"wednesday":2,"thu":3,"thursday":3,"fri":4,"friday":4,"sat":5,"saturday":5,"sun":6,"sunday":6}
def parse_semester(value:str)->int:
 value=value.strip().upper().replace("SEMESTER","").replace("SEM","").strip()
 roman={"I":1,"II":2,"III":3,"IV":4,"V":5,"VI":6,"VII":7,"VIII":8}
 return roman.get(value,int(value) if value.isdigit() else -1)
def parse_clock(value:str)->time:
 for pattern in ("%H:%M","%H:%M:%S"):
  try:return datetime.strptime(value.strip(),pattern).time()
  except ValueError:pass
 raise ValueError("time must use HH:MM")
def template_bytes(columns:list[str],rows:list[list[str]],xlsx:bool)->bytes:
 if not xlsx:return (",".join(columns)+"\n"+"\n".join(",".join(row) for row in rows)+"\n").encode()
 book=Workbook();sheet=book.active;sheet.title="Timetable";sheet.append(columns)
 for cell in sheet[1]:cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="1F4E78")
 for row in rows:sheet.append(row)
 sheet.freeze_panes="A2"
 for column in sheet.columns:sheet.column_dimensions[column[0].column_letter].width=max(14,min(32,max(len(str(x.value or "")) for x in column)+2))
 out=io.BytesIO();book.save(out);return out.getvalue()
def timetable_template(kind:str,xlsx:bool):
 columns=TEACHER_TIMETABLE_COLUMNS if kind=="teacher" else SECTION_ROUTINE_COLUMNS
 rows=[["NPT3F2509IT","SEM VI","A3|A4","SUN","08:30","09:30","CT004-3-3","Lecture","Block B","Machapuchare-L04"]] if kind=="teacher" else [["SUN","08:30","09:30","A3|A4","CT004-3-3","Advanced Database Systems","Lecture","karan@example.com","Block B","Machapuchare-L04"]]
 media="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if xlsx else "text/csv"
 return stream_bytes(template_bytes(columns,rows,xlsx),media,f"{kind}_timetable_template.{ 'xlsx' if xlsx else 'csv'}")
def import_error(row:int,field:str,message:str):return {"row_number":row,"field":field,"status":"invalid","error_message":message}
@dataclass
class RoutinePayloadResolution:
    payload: RoutineCreate
    pending_section_names: list[str]
    pending_reasons: dict[str, str]


def resolve_routine_payload_with_pending(db,row,teacher_id:int|None=None,context:tuple[int,int,int]|None=None):
 value=lambda key:str(row.get(key,"")).strip()
 day=DAY_CODES.get(value("day").lower() or value("day_of_week").lower())
 if day is None:raise ValueError("day: use MON through SUN")
 intake_code=value("intake_code")
 if not context and not intake_code:
  raise ValueError("This section routine worksheet must be imported from the Routine page after selecting its intake, semester, and section.")
 intake=db.scalar(select(Intake).where(func.lower(Intake.code)==intake_code.lower())) if not context else db.get(Intake,context[0])
 semester=parse_semester(value("semester") or value("semester_number")) if not context else context[1]
 if not intake:
  supplied=intake_code or (str(context[0]) if context else "")
  raise ValueError(f"Intake '{supplied}' does not exist. Select or create the intake before importing this routine.")
 if semester<1:raise ValueError("semester: enter a valid semester number")
 sections=[];seen_names=set()
 for part in (value("sections") or value("section_name")).split("|"):
  name=part.strip();key=name.casefold()
  if name and key not in seen_names:sections.append(name);seen_names.add(key)
 if not sections:raise ValueError("sections: provide at least one section name")
 selected_section=db.get(Section,context[2]) if context else None
 if context and not selected_section:raise ValueError("The selected import section no longer exists")
 section_rows={};pending_names=[];pending_reasons={}
 for name in sections:
  if context and name.casefold()==selected_section.name.casefold():
   section_rows[name.casefold()]=selected_section;continue
  candidates=db.scalars(select(Section).where(func.lower(Section.name)==name.lower(),or_(Section.intake_id==intake.id,Section.intake_id.is_(None)),or_(Section.semester_number==semester,Section.semester_number.is_(None)))).all()
  if not candidates:
   if context:
    pending_names.append(name);pending_reasons[name.casefold()]=f"Section {name} has not been configured for {intake.name} ({intake.code}), Semester {semester}."
    continue
   raise ValueError(f"No section named '{name}' is configured for {intake.name} ({intake.code}), Semester {semester}.")
  if len(candidates)>1:raise ValueError(f"sections: section {name} is ambiguous for the selected academic context")
  section_rows[name.casefold()]=candidates[0]
 if context and selected_section.name.casefold() not in section_rows:
  raise ValueError(f"Selected section '{selected_section.name}' is not included in this row.")
 module_code=value("module_code")
 module=db.scalar(select(AcademicModule).where(func.lower(AcademicModule.code)==module_code.lower()))
 class_type_name=value("class_type")
 class_type=db.scalar(select(ClassType).where(func.lower(ClassType.name)==class_type_name.lower()))
 lecturer_email=value("lecturer_email") or value("teacher_email")
 teacher=db.get(Teacher,teacher_id) if teacher_id else db.scalar(select(Teacher).join(User).where(func.lower(User.email)==lecturer_email.lower()))
 block_name=value("block") or value("block_name")
 block=db.scalar(select(Block).where(func.lower(Block.name)==block_name.lower()))
 room_name=value("room") or value("room_name")
 room=db.scalar(select(Room).where(Room.block_id==block.id if block else False,func.lower(Room.name)==room_name.lower()))
 try:start_time=parse_clock(value("start_time"));end_time=parse_clock(value("end_time"))
 except ValueError as exc:raise ValueError(f"start_time/end_time: {exc}") from exc
 slot=db.scalar(select(TimeSlot).where(TimeSlot.start_time==start_time,TimeSlot.end_time==end_time))
 if not module:raise ValueError(f"No Academic Module matches module code '{module_code}'. Create the module before importing this routine.")
 if module.semester_number!=semester:raise ValueError(f"Module '{module_code}' belongs to Semester {module.semester_number}, not Semester {semester}.")
 if value("module_title") and module.title.lower()!=value("module_title").lower():raise ValueError(f"Module title '{value('module_title')}' does not match module code '{module_code}'.")
 if not class_type:raise ValueError(f"Class type '{class_type_name}' does not exist. Create it before importing this routine.")
 if not teacher:raise ValueError(f"No teacher account/profile matches lecturer email '{lecturer_email}'. Create the teacher before importing this routine.")
 if not block:raise ValueError(f"Block '{block_name}' does not exist. Create the block before importing this routine.")
 if not room:raise ValueError(f"Room '{room_name}' does not exist in Block '{block_name}'.")
 if not slot:raise ValueError(f"No configured Time Slot exists for {start_time:%H:%M}-{end_time:%H:%M}. Create it before importing this routine.")
 if context:
  selected_payload=RoutineCreate(intake_id=intake.id,semester_number=semester,section_id=selected_section.id,section_ids=[selected_section.id],module_id=module.id,class_type_id=class_type.id,teacher_id=teacher.id,room_id=room.id,day_of_week=day,time_slot_id=slot.id)
  offering=valid_routine(db,selected_payload)
  offering_sections=offering_section_ids(db,offering);resolved=[selected_section]
  for name in sections:
   section=section_rows.get(name.casefold())
   if section is None or section.id==selected_section.id:continue
   if section.intake_id not in (None,intake.id) or section.semester_number not in (None,semester):
    pending_names.append(name);pending_reasons[name.casefold()]=f"Section {name} exists but is not configured for {intake.name} ({intake.code}), Semester {semester}.";continue
   if section.batch_id!=offering.batch_id:
    pending_names.append(name);pending_reasons[name.casefold()]=f"Section {name} exists but belongs to a different batch from the selected section.";continue
   if section.id not in offering_sections:
    pending_names.append(name);pending_reasons[name.casefold()]=f"Section {name} exists but is not included in the active Module Offering for {module.code}.";continue
   resolved.append(section)
  payload=selected_payload.model_copy(update={"section_ids":[section.id for section in resolved]})
  valid_routine(db,payload)
 else:
  resolved=list(section_rows.values())
  payload=RoutineCreate(intake_id=intake.id,semester_number=semester,section_id=resolved[0].id,section_ids=[section.id for section in resolved],module_id=module.id,class_type_id=class_type.id,teacher_id=teacher.id,room_id=room.id,day_of_week=day,time_slot_id=slot.id)
  valid_routine(db,payload)
 pending_names=list(dict.fromkeys(pending_names))
 return RoutinePayloadResolution(payload=payload,pending_section_names=pending_names,pending_reasons=pending_reasons)


def resolve_routine_payload(db,row,teacher_id:int|None=None,context:tuple[int,int,int]|None=None):
 resolution=resolve_routine_payload_with_pending(db,row,teacher_id,context)
 if resolution.pending_section_names:raise ValueError("Combined section references must be resolved before this import can proceed")
 return resolution.payload


def pending_warning_message(resolution:RoutinePayloadResolution)->str:
 return " Pending combined section(s): " + "; ".join(resolution.pending_reasons.get(name.casefold(),f"Section {name} is pending.") for name in resolution.pending_section_names) + " Create the section and include it in the active Module Offering before resolving."


async def validate_timetable_file(db,file,teacher_id:int|None=None,context:tuple[int,int,int]|None=None):
 frame=await read_import_file(file);valid=[];errors=[]
 for offset,row in frame.iterrows():
  try:
   resolution=resolve_routine_payload_with_pending(db,row,teacher_id,context);payload=resolution.payload;pending_names=resolution.pending_section_names;existing=matching_physical_routine(db,payload)
   if existing:
    additions=set(payload.section_ids or [payload.section_id])-set(link.section_id for link in existing.section_links)
    if additions:
     candidate=payload.model_copy(update={"section_id":next(iter(additions)),"section_ids":list(additions)});valid_routine(db,candidate);check_routine_conflicts(db,candidate,existing.id);state="valid_merge";message="Existing physical class found; new section membership(s) will be added."
    else:state="valid_existing";message="This physical class already exists and already includes these sections."
   else:
    check_routine_conflicts(db,payload);state="valid_new";message="New physical class is valid."
   if pending_names:
    state=f"{state}_with_pending";message+=pending_warning_message(resolution)
   valid.append((int(offset)+2,payload,state,message,pending_names))
  except Exception as exc:errors.append(import_error(int(offset)+2,"row",str(exc)))
 return len(frame),valid,errors
def read_job(job):return ImportJobRead(id=job.id,file_name=job.file_name,upload_type=job.upload_type,total_rows=job.total_rows,success_count=job.success_count,failed_count=job.failed_count,pending_section_references=job.pending_section_references,errors=[ImportError(**e) for e in json.loads(job.errors_json)],created_at=job.created_at)
async def read_import_file(file:UploadFile):
    raw=await file.read();name=(file.filename or "").lower()
    try:
        if name.endswith(".xlsx"):
            workbook=pd.ExcelFile(io.BytesIO(raw))
            if workbook.sheet_names!=["Timetable"]:raise HTTPException(400,"XLSX must contain exactly one worksheet named Timetable")
            return pd.read_excel(workbook,sheet_name="Timetable",dtype=str).fillna("")
        if name.endswith(".csv"):return pd.read_csv(io.BytesIO(raw),dtype=str).fillna("")
    except Exception as exc:raise HTTPException(400,"Invalid spreadsheet file") from exc
    raise HTTPException(400,"Only CSV and XLSX files are supported")
@router.post("/imports/students",response_model=ImportJobRead)
async def import_students(user:Annotated[User,Depends(require_role("admin"))],db:DbSession,file:UploadFile=File(...)):
    frame=await read_import_file(file)
    job=ImportJob(uploaded_by=user.id,file_name=file.filename or "students.csv",upload_type="students",total_rows=len(frame),success_count=0,failed_count=0);db.add(job);db.flush();errors=[]
    for offset,row in frame.iterrows():
        row_number=int(offset)+2
        try:
            with db.begin_nested():
                name=str(row.get("name","")).strip();email=str(row.get("email","")).strip().lower();batch_name=str(row.get("batch_name","")).strip();section_name=str(row.get("section_name","")).strip()
                if not name:raise ValueError("name is required")
                if not email:raise ValueError("email is required")
                if db.scalar(select(User).where(func.lower(User.email)==email)):raise ValueError("email is already in use")
                section=db.scalar(select(Section).join(Batch).where(func.lower(Batch.name)==batch_name.lower(),func.lower(Section.name)==section_name.lower()))
                if not section:raise ValueError("batch_name or section_name does not exist")
                if db.scalar(select(Student).where(func.lower(Student.email)==email)):raise ValueError("student email is already imported")
                account=User(name=name,email=email,password_hash=hash_password(secrets.token_urlsafe(32)),role=UserRole.STUDENT)
                db.add(account);db.flush()
                student=Student(user_id=account.id,section_id=section.id,roll_number=str(row.get("roll_number","")).strip() or f"IMP-{job.id}-{row_number}",name=name,email=email);db.add(student);db.flush();issue_student_invitation(db,student,account,welcome=True);phone=str(row.get("phone","")).strip()
                if phone:db.add(Guardian(name=f"Guardian of {name}",student_id=student.id,phone=phone))
            job.success_count+=1
        except Exception as exc:errors.append({"row_number":row_number,"error_message":str(exc)});job.failed_count+=1
    job.errors_json=json.dumps(errors);log_audit(db,user.id,"import.completed","import_job",job.id,None,{"success":job.success_count,"failed":job.failed_count});db.commit();db.refresh(job);return read_job(job)
@router.post("/imports/routines",response_model=ImportJobRead)
async def import_routines(user:Annotated[User,Depends(require_role("admin"))],db:DbSession,file:UploadFile=File(...)):
    frame=await read_import_file(file);job=ImportJob(uploaded_by=user.id,file_name=file.filename or "routines.csv",upload_type="routines",total_rows=len(frame),success_count=0,failed_count=0);db.add(job);db.flush();errors=[]
    for offset,row in frame.iterrows():
        row_number=int(offset)+2
        try:
            with db.begin_nested():
                # Accept both the legacy bulk columns and the canonical routine-sheet columns.
                # DAY_CODES supports MON/SUN as well as Monday/Sunday.
                payload=resolve_routine_payload(db,row)
                create_or_merge_routine_entry(db,payload)
            job.success_count+=1
        except Exception as exc:errors.append({"row_number":row_number,"error_message":str(exc)});job.failed_count+=1
    job.errors_json=json.dumps(errors);log_audit(db,user.id,"import.completed","import_job",job.id,None,{"upload_type":"routines","success":job.success_count,"failed":job.failed_count});db.commit();db.refresh(job);return read_job(job)
@router.get("/imports",response_model=list[ImportJobRead])
def import_history(user:Annotated[User,Depends(require_role("admin"))],db:DbSession):return [read_job(j) for j in db.scalars(select(ImportJob).order_by(ImportJob.created_at.desc())).all()]
@router.get("/imports/{id}",response_model=ImportJobRead)
def import_detail(id:int,user:Annotated[User,Depends(require_role("admin"))],db:DbSession):
    job=db.get(ImportJob,id)
    if not job:raise HTTPException(404,"Import job not found")
    return read_job(job)

@router.get("/academic/teachers/{teacher_id}/timetable/template")
def teacher_timetable_template(teacher_id:int,user:Annotated[User,Depends(require_role("admin"))],db:DbSession,format:str="csv"):
    if not db.get(Teacher,teacher_id):raise HTTPException(404,"Teacher not found")
    return timetable_template("teacher",format.lower()=="xlsx")
@router.get("/academic/sections/{section_id}/routine/template")
def section_routine_template(section_id:int,user:Annotated[User,Depends(require_role("admin"))],db:DbSession,format:str="csv"):
    if not db.get(Section,section_id):raise HTTPException(404,"Section not found")
    return timetable_template("section",format.lower()=="xlsx")
def preview_result(total,valid,errors):
 counts={"new":0,"existing":0,"merge":0,"pending":0}
 for _,_,state,_,pending_names in valid:
  if state.startswith("valid_new"):counts["new"]+=1
  elif state.startswith("valid_existing"):counts["existing"]+=1
  elif state.startswith("valid_merge"):counts["merge"]+=1
  counts["pending"]+=len(pending_names)
 return {"total_rows":total,"valid_rows":len(valid),"invalid_rows":len(errors),"new_rows":counts["new"],"existing_rows":counts["existing"],"merge_rows":counts["merge"],"pending_section_references":counts["pending"],"rows":[{"row":row,"status":state,"message":message,"pending_section_names":pending_names} for row,_,state,message,pending_names in valid],"errors":errors}
@router.post("/academic/teachers/{teacher_id}/timetable/preview")
async def preview_teacher_timetable(teacher_id:int,user:Annotated[User,Depends(require_role("admin"))],db:DbSession,file:UploadFile=File(...)):
    if not db.get(Teacher,teacher_id):raise HTTPException(404,"Teacher not found")
    total,valid,errors=await validate_timetable_file(db,file,teacher_id=teacher_id);return preview_result(total,valid,errors)
@router.post("/academic/sections/{section_id}/routine/preview")
async def preview_section_routine(section_id:int,intake_id:int,semester_number:int,user:Annotated[User,Depends(require_role("admin"))],db:DbSession,file:UploadFile=File(...)):
    section=db.get(Section,section_id)
    if not section:raise HTTPException(404,"Section not found")
    if section.intake_id!=intake_id or section.semester_number!=semester_number:raise HTTPException(422,"Selected section does not belong to selected intake and semester")
    total,valid,errors=await validate_timetable_file(db,file,context=(intake_id,semester_number,section_id));return preview_result(total,valid,errors)
@router.get("/academic/sections/{section_id}/routine/pending")
def pending_section_routine_references(section_id:int,intake_id:int,semester_number:int,user:Annotated[User,Depends(require_role("admin"))],db:DbSession):
    section=db.get(Section,section_id)
    if not section:raise HTTPException(404,"Section not found")
    if section.intake_id!=intake_id or section.semester_number!=semester_number:raise HTTPException(422,"Selected section does not belong to selected intake and semester")
    q=select(RoutinePendingSection).join(RoutineEntry).outerjoin(RoutineEntrySection).where(RoutinePendingSection.resolved_section_id.is_(None),RoutineEntry.intake_id==intake_id,RoutineEntry.semester_number==semester_number,or_(RoutineEntry.section_id==section_id,RoutineEntrySection.section_id==section_id)).order_by(RoutinePendingSection.created_at,RoutinePendingSection.id)
    return [{"id":item.id,"routine_entry_id":item.routine_entry_id,"section_name":item.section_name,"intake_id":item.intake_id,"semester_number":item.semester_number,"module_id":item.routine_entry.module_id,"module_code":item.routine_entry.module.code,"module_title":item.routine_entry.module.title,"day_of_week":item.routine_entry.day_of_week,"time_slot_id":item.routine_entry.time_slot_id,"teacher_id":item.routine_entry.teacher_id,"room_id":item.routine_entry.room_id,"class_type_id":item.routine_entry.class_type_id} for item in db.scalars(q).unique().all()]
def apply_timetable_import(db,user,file_name,total,valid,errors,upload_type):
    job=ImportJob(uploaded_by=user.id,file_name=file_name,upload_type=upload_type,total_rows=total,success_count=0,failed_count=len(errors),pending_section_references=0);db.add(job);db.flush()
    for row_number,payload,_,_,pending_names in valid:
        try:
            with db.begin_nested():
                entry,state=create_or_merge_routine_entry(db,payload)
                persist_pending_section_references(db,entry,pending_names,payload.intake_id,payload.semester_number)
            job.success_count+=1
            job.pending_section_references+=len(pending_names)
        except Exception as exc:job.failed_count+=1;errors.append(import_error(row_number,"row",str(exc)))
    job.errors_json=json.dumps(errors);log_audit(db,user.id,"import.completed","import_job",job.id,None,{"upload_type":upload_type,"success":job.success_count,"failed":job.failed_count,"pending_section_references":job.pending_section_references});db.commit();db.refresh(job);return read_job(job)
@router.post("/academic/teachers/{teacher_id}/timetable/import",response_model=ImportJobRead)
async def import_teacher_timetable(teacher_id:int,user:Annotated[User,Depends(require_role("admin"))],db:DbSession,file:UploadFile=File(...)):
    if not db.get(Teacher,teacher_id):raise HTTPException(404,"Teacher not found")
    total,valid,errors=await validate_timetable_file(db,file,teacher_id=teacher_id)
    return apply_timetable_import(db,user,file.filename or "teacher_timetable.csv",total,valid,errors,"teacher_timetable")
@router.post("/academic/sections/{section_id}/routine/import",response_model=ImportJobRead)
async def import_section_routine(section_id:int,intake_id:int,semester_number:int,user:Annotated[User,Depends(require_role("admin"))],db:DbSession,file:UploadFile=File(...)):
    section=db.get(Section,section_id)
    if not section:raise HTTPException(404,"Section not found")
    if section.intake_id!=intake_id or section.semester_number!=semester_number:raise HTTPException(422,"Selected section does not belong to selected intake and semester")
    total,valid,errors=await validate_timetable_file(db,file,context=(intake_id,semester_number,section_id))
    return apply_timetable_import(db,user,file.filename or "section_routine.csv",total,valid,errors,"section_routine")
@router.get("/academic/teachers/{teacher_id}/timetable/export")
def export_teacher_timetable(teacher_id:int,user:Annotated[User,Depends(require_role("admin"))],db:DbSession,format:str="csv"):
    teacher=db.get(Teacher,teacher_id)
    if not teacher:raise HTTPException(404,"Teacher not found")
    entries=db.scalars(select(RoutineEntry).where(RoutineEntry.teacher_id==teacher_id)).all()
    rows=[]
    for entry in entries:
        names="|".join(link.section.name for link in entry.section_links) or entry.section.name
        rows.append([entry.intake.code,str(entry.semester_number),names,["MON","TUE","WED","THU","FRI","SAT","SUN"][entry.day_of_week],str(entry.time_slot.start_time)[:5],str(entry.time_slot.end_time)[:5],entry.module.code,entry.class_type.name,entry.room.block.name,entry.room.name])
    xlsx=format.lower()=="xlsx";media="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if xlsx else "text/csv"
    return stream_bytes(template_bytes(TEACHER_TIMETABLE_COLUMNS,rows,xlsx),media,f"teacher_{teacher_id}_timetable.{ 'xlsx' if xlsx else 'csv'}")
def attendance_frame(db,user,student_id,section_id,batch_id,date_from,date_to):
    if date_from>date_to:raise HTTPException(422,"date_from must not exceed date_to")
    if user.role==UserRole.STUDENT:
        own=current_student_profile(db,user)
        if student_id and student_id!=own.id:raise HTTPException(403,"Students may only export their own attendance")
        student_id=own.id
    q=select(AttendanceRecord.id,Student.id.label("student_id"),User.name.label("student_name"),Student.roll_number,func.coalesce(AcademicModule.title,Subject.name).label("subject"),ClassSession.session_date,AttendanceRecord.status,AttendanceRecord.check_in_time).join(Student,AttendanceRecord.student_id==Student.id).join(User,Student.user_id==User.id).join(ClassSession,AttendanceRecord.class_session_id==ClassSession.id).outerjoin(TimetableEntry,ClassSession.timetable_entry_id==TimetableEntry.id).outerjoin(Subject,TimetableEntry.subject_id==Subject.id).outerjoin(RoutineEntry,ClassSession.routine_entry_id==RoutineEntry.id).outerjoin(AcademicModule,RoutineEntry.module_id==AcademicModule.id).join(Section,Student.section_id==Section.id).where(ClassSession.session_date.between(date_from,date_to))
    if user.role==UserRole.TEACHER:
        teacher=db.scalar(select(Teacher).where(Teacher.user_id==user.id))
        if not teacher:raise HTTPException(404,"Teacher profile not found")
        q=q.where(ClassSession.effective_teacher_id==teacher.id)
    if student_id:q=q.where(Student.id==student_id)
    if section_id:q=q.where(Student.section_id==section_id)
    if batch_id:q=q.where(Section.batch_id==batch_id)
    rows=db.execute(q.order_by(ClassSession.session_date,User.name)).mappings().all();frame=pd.DataFrame(rows)
    if len(frame):frame["status"]=frame["status"].map(lambda value:value.value if hasattr(value,"value") else str(value))
    return frame
def stream_bytes(data:bytes,media:str,filename:str):return StreamingResponse(io.BytesIO(data),media_type=media,headers={"Content-Disposition":f'attachment; filename="{filename}"'})
def render_pdf(html:str,title:str,lines:list[str],frame:pd.DataFrame)->bytes:
    try:
        from weasyprint import HTML
        return HTML(string=html).write_pdf()
    except Exception:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4,landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph,SimpleDocTemplate,Spacer,Table,TableStyle
        output=io.BytesIO();doc=SimpleDocTemplate(output,pagesize=landscape(A4));styles=getSampleStyleSheet();story=[Paragraph(settings.college_name,styles["Title"]),Paragraph(title,styles["Heading2"])]
        for line in lines:story.extend([Paragraph(line,styles["BodyText"]),Spacer(1,5)])
        data=[list(frame.columns)]+[[str(value) for value in row] for row in frame.itertuples(index=False,name=None)]
        if data:table=Table(data,repeatRows=1);table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#d1fae5")),("GRID",(0,0),(-1,-1),0.5,colors.grey),("FONTSIZE",(0,0),(-1,-1),7)]));story.append(table)
        doc.build(story);return output.getvalue()
@router.get("/exports/attendance.csv")
def attendance_csv(user:Annotated[User,Depends(get_current_user)],db:DbSession,date_from:date,date_to:date,student_id:int|None=None,section_id:int|None=None,batch_id:int|None=None):return stream_bytes(attendance_frame(db,user,student_id,section_id,batch_id,date_from,date_to).to_csv(index=False).encode(),"text/csv","attendance_report.csv")
@router.get("/exports/attendance.pdf")
def attendance_pdf(user:Annotated[User,Depends(get_current_user)],db:DbSession,date_from:date,date_to:date,student_id:int|None=None,section_id:int|None=None,batch_id:int|None=None):
    frame=attendance_frame(db,user,student_id,section_id,batch_id,date_from,date_to);passing=frame["status"].astype(str).str.lower().isin(["attendancestatus.present","attendancestatus.late","present","late"]).sum() if len(frame) else 0;percent=round(100*passing/len(frame),2) if len(frame) else 0;generated=datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC");html=Template("""<html><head><style>body{font-family:sans-serif}h1{color:#064e3b}table{width:100%;border-collapse:collapse}th,td{border:1px solid #aaa;padding:6px;font-size:11px}th{background:#d1fae5}</style></head><body><h1>{{college}}</h1><h2>Attendance Report</h2><p>Generated {{generated}} · {{date_from}} to {{date_to}}</p><p><b>Overall attendance: {{percent}}%</b> ({{passing}}/{{total}})</p>{{table|safe}}</body></html>""").render(college=settings.college_name,generated=generated,date_from=date_from,date_to=date_to,percent=percent,passing=passing,total=len(frame),table=frame.to_html(index=False));pdf=render_pdf(html,"Attendance Report",[f"Generated {generated}",f"Filters: {date_from} to {date_to}",f"Overall attendance: {percent}% ({passing}/{len(frame)})"],frame);return stream_bytes(pdf,"application/pdf","attendance_report.pdf")
def case_frame(db,status,date_from,date_to):
    q=select(StudentCase.id,StudentCase.student_id,StudentCase.trigger_type,StudentCase.scope_type,StudentCase.scope_id,StudentCase.status,StudentCase.priority,StudentCase.assigned_to,StudentCase.opened_at,StudentCase.closed_at)
    if status:q=q.where(StudentCase.status==CaseStatus(status))
    if date_from:q=q.where(func.date(StudentCase.opened_at)>=date_from)
    if date_to:q=q.where(func.date(StudentCase.opened_at)<=date_to)
    return pd.DataFrame(db.execute(q).mappings().all())
@router.get("/exports/cases.csv",dependencies=[Depends(require_role("admin"))])
def cases_csv(db:DbSession,status:str|None=None,date_from:date|None=None,date_to:date|None=None):return stream_bytes(case_frame(db,status,date_from,date_to).to_csv(index=False).encode(),"text/csv","case_report.csv")
@router.get("/exports/cases.pdf",dependencies=[Depends(require_role("admin"))])
def cases_pdf(db:DbSession,status:str|None=None,date_from:date|None=None,date_to:date|None=None):
    frame=case_frame(db,status,date_from,date_to);generated=f"{datetime.now(UTC):%Y-%m-%d}";html=f"<h1>{settings.college_name}</h1><h2>Student Case Report</h2><p>Generated {generated}</p>{frame.to_html(index=False)}";return stream_bytes(render_pdf(html,"Student Case Report",[f"Generated {generated}"],frame),"application/pdf","case_report.pdf")
@router.get("/exports/course-completion.csv",dependencies=[Depends(require_role("admin"))])
def course_csv(db:DbSession):
    rows=db.execute(select(CoursePlan.id,func.coalesce(AcademicModule.title,Subject.name).label("course"),Batch.name.label("batch"),CoursePlan.planned_sessions,CoursePlan.conducted_sessions).outerjoin(Subject,CoursePlan.subject_id==Subject.id).outerjoin(ModuleOffering,CoursePlan.module_offering_id==ModuleOffering.id).outerjoin(AcademicModule,ModuleOffering.academic_module_id==AcademicModule.id).join(Batch,CoursePlan.batch_id==Batch.id)).mappings().all();frame=pd.DataFrame(rows);frame["deficit"]=frame.planned_sessions-frame.conducted_sessions if len(frame) else [];return stream_bytes(frame.to_csv(index=False).encode(),"text/csv","course_completion.csv")
@router.get("/notifications",response_model=list[NotificationRead])
def notifications(user:Annotated[User,Depends(require_role("admin"))],db:DbSession,recipient_id:int|None=None,status:str|None=None):
    q=select(Notification)
    if recipient_id:q=q.where(Notification.recipient_id==recipient_id)
    if status:q=q.where(Notification.status==NotificationStatus(status))
    return db.scalars(q.order_by(Notification.created_at.desc())).all()
@router.get("/audit-logs",response_model=AuditPage)
def audit_logs(user:Annotated[User,Depends(require_role("admin"))],db:DbSession,actor_id:int|None=None,entity:str|None=None,date_from:date|None=None,date_to:date|None=None,page:int=1,page_size:int=50):
    q=select(AuditLog,User.name).join(User,AuditLog.actor_id==User.id)
    if actor_id:q=q.where(AuditLog.actor_id==actor_id)
    if entity:q=q.where(AuditLog.entity_type==entity)
    if date_from:q=q.where(func.date(AuditLog.created_at)>=date_from)
    if date_to:q=q.where(func.date(AuditLog.created_at)<=date_to)
    total=db.scalar(select(func.count()).select_from(q.subquery())) or 0;rows=db.execute(q.order_by(AuditLog.created_at.desc()).offset((page-1)*page_size).limit(min(page_size,100))).all();return AuditPage(items=[AuditRead(id=a.id,actor_id=a.actor_id,actor_name=n,action=a.action,entity_type=a.entity_type,entity_id=a.entity_id,details=a.details,created_at=a.created_at) for a,n in rows],total=total,page=page,page_size=page_size)
@router.get("/operations/health")
def health():return {"module":"operations","status":"ok"}
